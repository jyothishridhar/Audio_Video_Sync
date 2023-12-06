import streamlit as st
import moviepy.editor as mp
import numpy as np
from scipy.stats import skew
import pandas as pd
import requests
from openpyxl import Workbook
from openpyxl import load_workbook
import os

def download_file(url, dest_path):
    response = requests.get(url)
    with open(dest_path, 'wb') as file:
        file.write(response.content)

def sync_and_report(video_path, output_path_sync, output_path_unsync, output_path_report, delay_offset_unsync=0.5):
    try:
        # Load the video file
        video = mp.VideoFileClip(video_path)

        # Extract the audio and video tracks
        audio = video.audio

        # Get the frame rate and duration of the video
        video_fps = video.fps
        video_duration = video.duration

        # Resample the audio track to match the frame rate of the video track
        audio_resampled = audio.set_fps(video_fps)

        # Get the number of frames for the resampled audio track
        audio_frame_count = int(audio_resampled.duration * audio_resampled.fps)

        # Get the number of frames for the video track
        video_frame_count = int(video_duration * video_fps)

        # Calculate the delay values and corresponding time for each frame for sync
        video_delay_values_sync = []
        video_frame_times_sync = []
        audio_delay_values_sync = []
        audio_frame_times_sync = []

        for frame_number in range(video_frame_count):
            frame_time = frame_number / video_fps
            frame_audio = audio_resampled.subclip(frame_time, frame_time + 1 / video_fps)
            frame_video_delay = 1 / video_fps
            frame_audio_delay = frame_audio.duration
            video_delay_values_sync.append(frame_video_delay)
            video_frame_times_sync.append(frame_time)
            audio_delay_values_sync.append(frame_audio_delay)
            audio_frame_times_sync.append(frame_time)

        # Calculate the skewness of the audio delay values for sync
        audio_skewness_sync = skew(audio_delay_values_sync)

        # Create a DataFrame for synchronization report
        report_data_sync = {
            'Sync_Frame': list(range(1, video_frame_count + 1)),
            'Sync_Duration_video': video_delay_values_sync,
            'Sync_Time': video_frame_times_sync,
            'Sync_Duration_audio': audio_delay_values_sync,
            'Sync_Summary': [sum(abs(video_delay - audio_delay) for video_delay, audio_delay in zip(video_delay_values_sync, audio_delay_values_sync))]
        }

        # Make sure all columns have the same length
        max_length = max(len(report_data_sync[col]) for col in report_data_sync)
        for col in report_data_sync:
            report_data_sync[col] += [np.nan] * (max_length - len(report_data_sync[col]))

        report_df_sync = pd.DataFrame(report_data_sync)

        # Create a Pandas Excel writer for the report
        excel_writer = pd.ExcelWriter(output_path_report, engine='openpyxl')

        # Write the DataFrame to the Excel file for synchronization report
        report_df_sync.to_excel(excel_writer, sheet_name='Sync_Frames', index=False)

        # Save the synchronization Excel file
        excel_writer.close()

        # Open the existing workbook
        book = load_workbook(output_path_report)

        # Access the active sheet
        worksheet_sync = book['Sync_Frames']

        # Write additional information to the Excel file for synchronization report
        worksheet_sync.cell(row=video_frame_count + 3, column=1, value='Audio_Video_delay_seconds:')
        worksheet_sync.cell(row=video_frame_count + 3, column=2, value=report_data_sync['Sync_Summary'][0])
        worksheet_sync.cell(row=video_frame_count + 6, column=1, value='Audio Skewness:')
        worksheet_sync.cell(row=video_frame_count + 6, column=2, value=audio_skewness_sync)
        worksheet_sync.cell(row=video_frame_count + 8, column=1, value='Number of frames in audio:')
        worksheet_sync.cell(row=video_frame_count + 8, column=2, value=audio_frame_count)
        worksheet_sync.cell(row=video_frame_count + 9, column=1, value='Number of frames in video:')
        worksheet_sync.cell(row=video_frame_count + 9, column=2, value=video_frame_count)

        # Save the synchronization Excel file
        book.save(output_path_report)

        # Write the synchronized video to the output file
        synced_video = video.set_audio(audio)
        synced_video.write_videofile(output_path_sync, codec='libx264', audio_codec='aac')

        # Load the synchronized video file
        video_synced = mp.VideoFileClip(output_path_sync)
        audio_synced = video_synced.audio

        # Calculate the delay values and corresponding time for each frame for unsync
        video_delay_values_unsync = []
        video_frame_times_unsync = []
        audio_delay_values_unsync = []
        audio_frame_times_unsync = []

        for frame_number in range(video_frame_count):
            frame_time = frame_number / video_fps
            t_start = frame_time + delay_offset_unsync
            t_end = frame_time + 1 / video_fps + delay_offset_unsync

            if t_start <= video_duration:
                frame_audio = audio_resampled.subclip(max(t_start, 0), min(t_end, video_duration))
                frame_video_delay = 1 / video_fps
                frame_audio_delay = frame_audio.duration
                video_delay_values_unsync.append(frame_video_delay)
                video_frame_times_unsync.append(frame_time)
                audio_delay_values_unsync.append(frame_audio_delay)
                audio_frame_times_unsync.append(frame_time)

        # Calculate the total delay for video and audio separately for unsync
        total_video_delay_unsync = sum(video_delay_values_unsync)
        total_audio_delay_unsync = sum(audio_delay_values_unsync)

        # Calculate the skewness of the audio delay values for unsync
        audio_skewness_unsync = skew(audio_delay_values_unsync)

        # Create a DataFrame with the delay and time for each video frame for unsync
        report_data_unsync = {
            'Unsync_Frame': list(range(1, len(video_delay_values_unsync) + 1)),
            'Unsync_Duration_video': video_delay_values_unsync,
            'Unsync_Time': video_frame_times_unsync,
            'Unsync_Duration_audio': audio_delay_values_unsync,
            'Unsync_Summary': [total_video_delay_unsync]
        }

        # Make sure all columns have the same length
        max_length_unsync = max(len(report_data_unsync[col]) for col in report_data_unsync)
        for col in report_data_unsync:
            report_data_unsync[col] += [np.nan] * (max_length_unsync - len(report_data_unsync[col]))

        report_df_unsync = pd.DataFrame(report_data_unsync)

        # Add unsync data to the existing Excel file
        with pd.ExcelWriter(output_path_report, engine='openpyxl', mode='a') as writer:
            # Write the DataFrame to the Excel file for unsynchronization report
            report_df_unsync.to_excel(writer, sheet_name='Unsync_Frames', index=False)

            # Write additional information to the Excel file for unsynchronization report
            worksheet_unsync = writer.sheets['Unsync_Frames']
            worksheet_unsync.cell(row=video_frame_count + 4, column=1, value='Audio_Video_delay_seconds:')
            worksheet_unsync.cell(row=video_frame_count + 4, column=2, value=total_video_delay_unsync)
            worksheet_unsync.cell(row=video_frame_count + 7, column=1, value='Audio Skewness:')
            worksheet_unsync.cell(row=video_frame_count + 7, column=2, value=audio_skewness_unsync)
            worksheet_unsync.cell(row=video_frame_count + 9, column=1, value='Number of frames in audio:')
            worksheet_unsync.cell(row=video_frame_count + 9, column=2, value=audio_frame_count)
            worksheet_unsync.cell(row=video_frame_count + 10, column=1, value='Number of frames in video:')
            worksheet_unsync.cell(row=video_frame_count + 10, column=2, value=video_frame_count)

        # Display download links
        st.markdown(f"Download [Synchronized Video]({output_path_sync})")
        st.markdown(f"Download [Unsynchronized Video]({output_path_unsync})")
        st.markdown(f"Download [Synchronization Report]({output_path_report})")

        st.success("Processing completed successfully!")

    except Exception as e:
        st.error(f"Error during processing: {e}")
        # Print the exception traceback for detailed error information
        import traceback
        st.text(traceback.format_exc())

        # Clean up the temporary video files even if an error occurs
        if os.path.exists(video_path):
            os.unlink(video_path)
        if os.path.exists(output_path_sync):
            os.unlink(output_path_sync)
        if os.path.exists(output_path_unsync):
            os.unlink(output_path_unsync)

# Streamlit UI
st.title("Audio-Video Synchronization App")

# File uploader
uploaded_file = st.file_uploader("Upload a video file", type=["mp4", "avi"])

if uploaded_file is not None:
    # Temporary file paths
    video_dest_path = "video.mp4"
    output_path_sync = 'sync_video.mp4'
    output_path_unsync = 'unsync_video.mp4'
    output_path_report = 'report_combined.xlsx'

    # Save the uploaded file to a temporary location
    with open(video_dest_path, 'wb') as f:
        f.write(uploaded_file.read())

    # Perform synchronization and generate report
    sync_and_report(video_dest_path, output_path_sync, output_path_unsync, output_path_report)

    # Display download links
    st.markdown(f"Download [Synchronized Video]({output_path_sync})")
    st.markdown(f"Download [Unsynchronized Video]({output_path_unsync})")
    st.markdown(f"Download [Synchronization Report]({output_path_report})")
